# -*- coding: utf-8 -*-
# @Time :    2020/11/21 17:33
# @Author :  Dewei
# @FileName: API_2.py
# @Email :   969182775@qq.com
# @IDE:      PyCharm
import openpyxl
import requests
# 读取Excel信息
def read_data(fileName,sheetName):
    wb = openpyxl.load_workbook(fileName,sheetName)
    sheet = wb[sheetName]
    max_row = sheet.max_row
    list_1 = []
    for i in range(2,max_row + 1):
        dict_1 = dict(
        id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,
        data = sheet.cell(row=i,column=6).value,
        expect = sheet.cell(row=i,column=7).value
                    )
        list_1.append(dict_1)
    return list_1

# 发送请求
def api_request(url,data):
    header ={"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}#传请求头
    res_1 = requests.post(url=url,json=data,headers=header)#定义一个变量接收request
    respones = res_1.json() #返回的结果是相应状态码，所以用.json格式
    return respones #返回结果

def token(url,data):
    tt = api_request(url,data)
    token = tt.json()['data']['token_info']['token']
    print("token:",token)
token()

# 写入断言结果
def write_res(fileName,sheetName,row,column,final_result):
    wb = openpyxl.load_workbook(fileName)
    sheet = wb[sheetName]
    sheet.cell(row=row,column=column).value = final_result
    wb.save(fileName)

#执行自动化测试用例
def run_exe(fileName,sheetName):
    cases = read_data(fileName,sheetName)
    for case in cases:
        id = case.get('id')
        url = case.get('url')
        data =case.get('data')
        expect = case.get('expect')
        expect = eval(expect)
        expect_msg = expect.get('msg')
        data = eval(data)
        real_result = api_request(url,data)
        real_msg = real_result.get('msg')
        print('执行结果为：{}'.format(real_msg))
        print('预期结果为：{}'.format(expect_msg))
        if expect_msg == real_result:
            print('这条测试用例通过！！！')
            final_result = 'pass'
        else:
            print('这条测试用例不通过！！！')
            final_result = 'fail'
        print('*' * 30)
        write_res(fileName,sheetName,id+1,8,final_result)
run_exe('test_case_api.xlsx','login')